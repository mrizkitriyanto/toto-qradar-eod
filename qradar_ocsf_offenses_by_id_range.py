#!/usr/bin/env python3
"""
Reference:
  GET /siem/offenses_ocsf -- QRadar API 28.0

  python3 qradar_ocsf_offenses_by_id_range.py --host qradar.example.com --start-id 5001 --end-id 5005 -u user
  python3 qradar_ocsf_offenses_by_id_range.py --host qradar.example.com --start-id 5001 --end-id 5005 -u user --insecure
  python3 qradar_ocsf_offenses_by_id_range.py --host qradar.example.com --start-id 5001 --end-id 5005 -u user --insecure --output-xlsx
  python3 qradar_ocsf_offenses_by_id_range.py --host qradar.example.com --start-id 5001 --end-id 5999 -u user --insecure --severity High,Critical --output-xlsx

"""

import argparse
import base64
import getpass
import json
import os
import re
import sqlite3
import ssl
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone

DEFAULT_API_VERSION = "28.0"
DEFAULT_TIMEOUT = 30
DEFAULT_PAGE_SIZE = 1000
DEFAULT_RETRIES = 3
DEFAULT_BACKOFF_SECONDS = 2
DEFAULT_BUFFER = 50          # positions padded around the estimated window
DEFAULT_MAX_EXPAND = 5       # how many times the buffer is doubled before giving up
MAX_PAGES_SAFETY_CAP = 10_000
LARGE_REQUEST_THRESHOLD = 500  # confirm before fetching more offense IDs than this

REPORT_TIMEZONE = timezone(timedelta(hours=7))  # Asia/Jakarta (WIB)
NA = "N/A"

DEFAULT_XLSX_NAME_PATTERN = "Daily_QRadar_ddmmyyyy-hh-mm-ss.xlsx"
DEFAULT_KB_DB_NAME = "kb.sqlite3"

EXCEL_COLUMNS = [
    "Offense ID",
    "Offense Name",
    "Severity",
    "Log Source",
    "Src IP",
    "Dst IP",
    "Offense Source",
    "Username",
    "Deskripsi",
    "Rekomendasi",
    "Action",
    "Timestamp",
]


# --------------------------------------------------------------------------
# Auth
# --------------------------------------------------------------------------

def build_basic_auth_header(username: str, password: str) -> str:
    token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
    return f"Basic {token}"


# --------------------------------------------------------------------------
# HTTP
# --------------------------------------------------------------------------

def http_get_range(host: str, range_header: str, headers: dict, ssl_context,
                    timeout: int, retries: int) -> tuple[list, dict]:
    """
    GET /api/siem/offenses_ocsf with a given Range header.
    Returns (parsed_json_list, lowercase_response_headers).
    """
    url = f"https://{host}/api/siem/offenses_ocsf"
    req_headers = dict(headers)
    req_headers["Range"] = range_header

    attempt = 0
    while True:
        attempt += 1
        req = urllib.request.Request(url, headers=req_headers, method="GET")
        try:
            with urllib.request.urlopen(req, timeout=timeout, context=ssl_context) as resp:
                raw = resp.read().decode("utf-8")
                data = json.loads(raw) if raw.strip() else []
                resp_headers = {k.lower(): v for k, v in resp.headers.items()}
                return data, resp_headers
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            if e.code == 401:
                raise RuntimeError("401 Unauthorized: check --username/--password.") from None
            if e.code == 403:
                raise RuntimeError(
                    "403 Forbidden: account lacks permission for offenses_ocsf."
                ) from None
            if e.code >= 500 and attempt <= retries:
                wait = DEFAULT_BACKOFF_SECONDS * (2 ** (attempt - 1))
                print(f"[WARN] HTTP {e.code}, retrying in {wait}s "
                      f"(attempt {attempt}/{retries})...", file=sys.stderr)
                time.sleep(wait)
                continue
            raise RuntimeError(f"HTTP {e.code} error: {body}") from None
        except urllib.error.URLError as e:
            if attempt <= retries:
                wait = DEFAULT_BACKOFF_SECONDS * (2 ** (attempt - 1))
                print(f"[WARN] Connection error ({e.reason}), retrying in {wait}s "
                      f"(attempt {attempt}/{retries})...", file=sys.stderr)
                time.sleep(wait)
                continue
            raise RuntimeError(f"Connection failed: {e.reason}") from None


def parse_content_range_total(resp_headers: dict):
    """Parse 'content-range: items 0-1/3882' -> 3882. Returns None if absent/unparseable."""
    cr = resp_headers.get("content-range")
    if not cr:
        return None
    m = re.search(r"/(\d+)\s*$", cr)
    return int(m.group(1)) if m else None


def get_uid(record: dict):
    """Safely extract finding_info.uid as int. Returns None if missing/malformed."""
    try:
        return int(record["finding_info"]["uid"])
    except (KeyError, TypeError, ValueError):
        return None


def fetch_positions(host: str, lo: int, hi: int, headers: dict, ssl_context,
                     timeout: int, retries: int, page_size: int) -> tuple[list, bool]:
    """
    Fetch offenses at list positions [lo, hi] inclusive, paging in page_size chunks.
    Returns (records, hit_end_of_list) -- hit_end_of_list is True if the server
    returned fewer items than requested at some point (i.e. we reached the end
    of the offense list before reaching `hi`).
    """
    results = []
    pos = max(lo, 0)
    hit_end = False
    pages = 0
    while pos <= hi and pages < MAX_PAGES_SAFETY_CAP:
        pages += 1
        end = min(pos + page_size - 1, hi)
        batch, _ = http_get_range(host, f"items={pos}-{end}", headers, ssl_context,
                                   timeout, retries)
        if not batch:
            hit_end = True
            break
        results.extend(batch)
        if len(batch) < (end - pos + 1):
            hit_end = True
            break
        pos = end + 1
    return results, hit_end


# --------------------------------------------------------------------------
# Strategy 1: smart position estimate + verification
# --------------------------------------------------------------------------

def smart_fetch(host: str, start_id: int, end_id: int, headers: dict, ssl_context,
                 timeout: int, retries: int, page_size: int, buffer: int,
                 max_expand: int) -> tuple[list, bool]:
    """
    Returns (matching_records, verified). `verified` is False if the position
    estimate could not be confirmed complete within the retry budget -- caller
    should fall back to full_scan() in that case.
    """
    probe, probe_headers = http_get_range(host, "items=0-1", headers, ssl_context,
                                           timeout, retries)
    if not probe:
        return [], True  # no offenses in the system at all

    uid0 = get_uid(probe[0])
    if uid0 is None:
        print("[WARN] Could not read finding_info.uid from the first offense; "
              "falling back to full scan.", file=sys.stderr)
        return [], False

    total_count = parse_content_range_total(probe_headers)

    if len(probe) >= 2:
        uid1 = get_uid(probe[1])
        descending = (uid1 is None) or (uid0 > uid1)
    else:
        descending = True  # only one offense exists; direction is moot

    if descending:
        est_lo = uid0 - end_id
        est_hi = uid0 - start_id
    else:
        est_lo = start_id - uid0
        est_hi = end_id - uid0

    attempt = 0
    cur_buffer = buffer
    last_batch, last_lo, last_hi, last_hit_end = [], 0, 0, False

    while attempt <= max_expand:
        lo = max(0, est_lo - cur_buffer)
        hi = est_hi + cur_buffer
        if total_count is not None:
            hi = min(hi, total_count - 1)

        batch, hit_end = fetch_positions(host, lo, hi, headers, ssl_context,
                                          timeout, retries, page_size)
        last_batch, last_lo, last_hi, last_hit_end = batch, lo, hi, hit_end

        if verify_window(batch, lo, hit_end, start_id, end_id, descending):
            matches = [r for r in batch if (u := get_uid(r)) is not None
                       and start_id <= u <= end_id]
            return matches, True

        attempt += 1
        cur_buffer *= 2
        if attempt <= max_expand:
            print(f"[INFO] Position estimate window looked incomplete, "
                  f"expanding buffer to {cur_buffer} and retrying "
                  f"({attempt}/{max_expand})...", file=sys.stderr)

    matches = [r for r in last_batch if (u := get_uid(r)) is not None
               and start_id <= u <= end_id]
    return matches, False


def verify_window(batch: list, lo: int, hit_end: bool, start_id: int, end_id: int,
                   descending: bool) -> bool:
    """
    Confirm the fetched window's edges are already outside the target ID range,
    which means no matching offense could exist just beyond what we fetched.
    """
    if not batch:
        return lo == 0 or hit_end  # nothing fetched; only OK if there was nothing to fetch

    ids = [u for r in batch if (u := get_uid(r)) is not None]
    if not ids:
        return False

    if descending:
        high_edge = ids[0]     # earliest position in window = largest ID in window
        low_edge = ids[-1]     # latest position in window = smallest ID in window
        high_ok = (lo == 0) or (high_edge > end_id)
        low_ok = hit_end or (low_edge < start_id)
    else:
        low_edge = ids[0]
        high_edge = ids[-1]
        low_ok = (lo == 0) or (low_edge < start_id)
        high_ok = hit_end or (high_edge > end_id)

    return high_ok and low_ok


# --------------------------------------------------------------------------
# Strategy 2: full linear scan (guaranteed correct, slower)
# --------------------------------------------------------------------------

def full_scan(host: str, start_id: int, end_id: int, headers: dict, ssl_context,
              timeout: int, retries: int, page_size: int) -> list:
    matches = []
    pos = 0
    pages = 0
    while pages < MAX_PAGES_SAFETY_CAP:
        pages += 1
        batch, resp_headers = http_get_range(
            host, f"items={pos}-{pos + page_size - 1}", headers, ssl_context,
            timeout, retries,
        )
        if not batch:
            break
        for r in batch:
            u = get_uid(r)
            if u is not None and start_id <= u <= end_id:
                matches.append(r)

        total_count = parse_content_range_total(resp_headers)
        if total_count is not None:
            print(f"[INFO] Full scan progress: {min(pos + page_size, total_count)}/"
                  f"{total_count}", file=sys.stderr)

        if len(batch) < page_size:
            break
        pos += page_size
    return matches


# --------------------------------------------------------------------------
# Excel export
#
# See "QRadar Offense Report - Excel Output Spec.md" for the full field
# mapping rationale. transform_to_rows() is the only place that knows how
# OCSF fields map to report columns; write_xlsx() just renders rows.
# --------------------------------------------------------------------------

def clean_text(value) -> str:
    """Collapse embedded newlines (offense titles/descs are multi-line rule
    chains) into a single-line, Excel-cell-friendly string."""
    if value is None:
        return NA
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = " ".join(line.strip() for line in text.split("\n") if line.strip())
    return text or NA


def get_observable_values(record: dict, name: str) -> list:
    return [
        o.get("value") for o in record.get("observables") or []
        if o.get("name") == name and o.get("value")
    ]


def get_endpoint_ips(record: dict, endpoint_key: str) -> list:
    ips = []
    for ev in record.get("evidences") or []:
        endpoint = ev.get(endpoint_key)
        if endpoint and endpoint.get("ip"):
            ips.append(endpoint["ip"])
    return ips


def get_ip_column(record: dict, observable_name: str, endpoint_key: str) -> str:
    ips = get_observable_values(record, observable_name) or get_endpoint_ips(record, endpoint_key)
    seen = list(dict.fromkeys(ips))  # de-dupe, preserve order
    return ", ".join(seen) if seen else NA


def get_dst_column(record: dict) -> str:
    """When the offense has remote destinations (enrichments.remote_destination_count
    > 0), report "Remote(<count>)" instead of the raw destination IP list."""
    remote_count = get_enrichment_int(record, "remote_destination_count")
    if remote_count > 0:
        return f"Remote({remote_count})"
    return get_ip_column(record, "dst_ip_address", "dst_endpoint")


def get_log_source(record: dict) -> str:
    """Log Source, excluding the CRE logger itself (log_provider 'EventCRE',
    e.g. 'Custom Rule Engine-8 :: sdcplxsiqap01') -- it's the correlation
    engine that fired the offense, not the originating log source."""
    names = [
        lg.get("log_name") for lg in (record.get("metadata") or {}).get("loggers") or []
        if lg.get("log_name")
        and lg.get("log_provider") != "EventCRE"
        and not lg["log_name"].strip().lower().startswith("custom rule engine")
    ]
    return "; ".join(dict.fromkeys(names)) if names else NA


def get_enrichment_value(record: dict, name: str):
    for e in record.get("enrichments") or []:
        if e.get("name") == name:
            return e.get("value")
    return None


def get_enrichment_int(record: dict, name: str) -> int:
    try:
        return int(get_enrichment_value(record, name))
    except (TypeError, ValueError):
        return 0


def get_username(record: dict) -> str:
    """Parse 'BY_USERNAME::<value>' out of metadata.correlation_uid. Other
    correlation types (BY_SOURCE_IP, etc.) have no username to report."""
    correlation_uid = (record.get("metadata") or {}).get("correlation_uid") or ""
    if "::" not in correlation_uid:
        return NA
    corr_type, _, value = correlation_uid.partition("::")
    if corr_type.strip().upper() == "BY_USERNAME" and value.strip():
        return value.strip()
    return NA


def format_timestamp_range(record: dict) -> str:
    def fmt(epoch_ms):
        if epoch_ms is None:
            return "?"
        dt = datetime.fromtimestamp(epoch_ms / 1000, tz=timezone.utc).astimezone(REPORT_TIMEZONE)
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    return f"{fmt(record.get('start_time'))} - {fmt(record.get('end_time'))}"


def get_severity(record: dict) -> str:
    value = record.get("severity") or record.get("severity_id")
    return str(value).strip() if value not in (None, "") else NA


def parse_severity_filter(raw: str) -> set:
    return {s.strip().lower() for s in raw.split(",") if s.strip()}


def filter_by_severity(matches: list, allowed: set) -> list:
    return [r for r in matches if get_severity(r).lower() in allowed]


# --------------------------------------------------------------------------
# KB (SQLite) knowledge base -- Deskripsi/Rekomendasi lookup
#
# Matching: an offense's finding_info.title is looked up against KB entries
# whose normalized keyword is a *substring* of that title; when several
# entries match (common, since titles are "X preceded by Y containing Z"
# rule chains), the longest/most specific keyword wins. See "QRadar Offense
# Report - Excel Output Spec.md" for the rationale and known limitations.
# --------------------------------------------------------------------------

def normalize_for_match(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def resolve_kb_db_path(raw_value) -> str:
    if raw_value:
        return raw_value
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, DEFAULT_KB_DB_NAME)


def kb_connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS kb_entries (
            id INTEGER PRIMARY KEY,
            offense_name TEXT NOT NULL,
            normalized_keyword TEXT NOT NULL UNIQUE,
            deskripsi TEXT NOT NULL,
            rekomendasi TEXT NOT NULL
        )
    """)
    conn.commit()
    return conn


def kb_entry_count(conn: sqlite3.Connection) -> int:
    return conn.execute("SELECT COUNT(*) FROM kb_entries").fetchone()[0]


def kb_lookup(conn: sqlite3.Connection, raw_title: str):
    """Return the sqlite3.Row (offense_name, deskripsi, rekomendasi) whose
    normalized keyword is the longest substring match of raw_title, or None."""
    normalized_title = normalize_for_match(raw_title)
    if not normalized_title:
        return None
    return conn.execute("""
        SELECT offense_name, deskripsi, rekomendasi
        FROM kb_entries
        WHERE instr(?, normalized_keyword) > 0
        ORDER BY length(normalized_keyword) DESC
        LIMIT 1
    """, (normalized_title,)).fetchone()


def kb_upsert(conn: sqlite3.Connection, keyword: str, deskripsi: str, rekomendasi: str,
              entry_id=None) -> int:
    normalized = normalize_for_match(keyword)
    if entry_id is not None:
        conn.execute("""
            INSERT INTO kb_entries (id, offense_name, normalized_keyword, deskripsi, rekomendasi)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(normalized_keyword) DO UPDATE SET
                offense_name = excluded.offense_name,
                deskripsi = excluded.deskripsi,
                rekomendasi = excluded.rekomendasi
        """, (entry_id, keyword, normalized, deskripsi, rekomendasi))
    else:
        conn.execute("""
            INSERT INTO kb_entries (offense_name, normalized_keyword, deskripsi, rekomendasi)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(normalized_keyword) DO UPDATE SET
                offense_name = excluded.offense_name,
                deskripsi = excluded.deskripsi,
                rekomendasi = excluded.rekomendasi
        """, (keyword, normalized, deskripsi, rekomendasi))
    conn.commit()
    row = conn.execute("SELECT id FROM kb_entries WHERE normalized_keyword = ?",
                        (normalized,)).fetchone()
    return row["id"]


def kb_import_xlsx(conn: sqlite3.Connection, xlsx_path: str) -> tuple:
    """Bulk-import rows (ID, Offense Name, Deskripsi, Rekomendasi) from an
    Excel KB file. Returns (inserted, updated). Idempotent: re-importing the
    same file updates existing rows instead of duplicating/erroring."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "--kb-import requires the 'openpyxl' package. Install it with: "
            "pip install openpyxl"
        ) from None

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    inserted, updated = 0, 0
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:  # skip header
        if not row or row[1] is None:
            continue
        entry_id, offense_name, deskripsi, rekomendasi = row[0], row[1], row[2], row[3]
        normalized = normalize_for_match(offense_name)
        existing = conn.execute("SELECT 1 FROM kb_entries WHERE normalized_keyword = ?",
                                 (normalized,)).fetchone()
        kb_upsert(conn, offense_name, deskripsi or NA, rekomendasi or NA, entry_id=entry_id)
        if existing:
            updated += 1
        else:
            inserted += 1
    return inserted, updated


def kb_update_entry(conn: sqlite3.Connection, entry_id: int, keyword=None, deskripsi=None,
                     rekomendasi=None) -> bool:
    row = conn.execute("SELECT * FROM kb_entries WHERE id = ?", (entry_id,)).fetchone()
    if row is None:
        return False
    new_keyword = keyword if keyword is not None else row["offense_name"]
    new_deskripsi = deskripsi if deskripsi is not None else row["deskripsi"]
    new_rekomendasi = rekomendasi if rekomendasi is not None else row["rekomendasi"]
    conn.execute("""
        UPDATE kb_entries
        SET offense_name = ?, normalized_keyword = ?, deskripsi = ?, rekomendasi = ?
        WHERE id = ?
    """, (new_keyword, normalize_for_match(new_keyword), new_deskripsi, new_rekomendasi, entry_id))
    conn.commit()
    return True


def kb_delete_entry(conn: sqlite3.Connection, entry_id: int) -> bool:
    cur = conn.execute("DELETE FROM kb_entries WHERE id = ?", (entry_id,))
    conn.commit()
    return cur.rowcount > 0


def kb_list_entries(conn: sqlite3.Connection) -> list:
    return conn.execute("SELECT id, offense_name FROM kb_entries ORDER BY id").fetchall()


def print_kb_list(conn: sqlite3.Connection) -> None:
    rows = kb_list_entries(conn)
    if not rows:
        print("[INFO] KB is empty.", file=sys.stderr)
        return
    for row in rows:
        name = row["offense_name"]
        if len(name) > 90:
            name = name[:87] + "..."
        print(f"{row['id']:>5}  {name}")


def run_kb_management(args, kb_db_path: str) -> None:
    conn = kb_connect(kb_db_path)
    try:
        if args.kb_import:
            inserted, updated = kb_import_xlsx(conn, args.kb_import)
            print(f"[INFO] KB import from {args.kb_import}: {inserted} inserted, "
                  f"{updated} updated.", file=sys.stderr)
        elif args.kb_list:
            print_kb_list(conn)
        elif args.kb_add:
            if not (args.kb_keyword and args.kb_deskripsi and args.kb_rekomendasi):
                print("[ERROR] --kb-add requires --kb-keyword, --kb-deskripsi, and "
                      "--kb-rekomendasi.", file=sys.stderr)
                sys.exit(2)
            entry_id = kb_upsert(conn, args.kb_keyword, args.kb_deskripsi, args.kb_rekomendasi)
            print(f"[INFO] KB entry saved (id={entry_id}).", file=sys.stderr)
        elif args.kb_update is not None:
            if not any([args.kb_keyword, args.kb_deskripsi, args.kb_rekomendasi]):
                print("[ERROR] --kb-update requires at least one of --kb-keyword/"
                      "--kb-deskripsi/--kb-rekomendasi.", file=sys.stderr)
                sys.exit(2)
            ok = kb_update_entry(conn, args.kb_update, args.kb_keyword, args.kb_deskripsi,
                                  args.kb_rekomendasi)
            if ok:
                print(f"[INFO] KB entry {args.kb_update} updated.", file=sys.stderr)
            else:
                print(f"[ERROR] KB entry {args.kb_update} not found.", file=sys.stderr)
                sys.exit(1)
        elif args.kb_delete is not None:
            ok = kb_delete_entry(conn, args.kb_delete)
            if ok:
                print(f"[INFO] KB entry {args.kb_delete} deleted.", file=sys.stderr)
            else:
                print(f"[ERROR] KB entry {args.kb_delete} not found.", file=sys.stderr)
                sys.exit(1)
    except RuntimeError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()


def transform_to_rows(matches: list, kb_conn: sqlite3.Connection = None) -> list:
    """Map each OCSF offense record to one report row, keyed by EXCEL_COLUMNS."""
    rows = []
    for r in matches:
        offense_source = get_enrichment_value(r, "offense_source")
        kb_hit = kb_lookup(kb_conn, r.get("finding_info", {}).get("title")) if kb_conn else None
        rows.append({
            "Offense ID": get_uid(r) if get_uid(r) is not None else NA,
            "Offense Name": clean_text(r.get("finding_info", {}).get("title")),
            "Severity": get_severity(r),
            "Log Source": get_log_source(r),
            "Src IP": get_ip_column(r, "src_ip_address", "src_endpoint"),
            "Dst IP": get_dst_column(r),
            "Offense Source": offense_source or NA,
            "Username": get_username(r),
            # Deskripsi/Rekomendasi come from the SQLite KB (kb_lookup, matched
            # by keyword-in-title with longest-match tie-break) when available.
            # Fallback when there's no KB hit (or KB disabled via --no-kb):
            # Deskripsi from finding_info.desc, Rekomendasi stays N/A.
            "Deskripsi": kb_hit["deskripsi"] if kb_hit else clean_text(r.get("finding_info", {}).get("desc")),
            "Rekomendasi": kb_hit["rekomendasi"] if kb_hit else NA,
            "Action": clean_text(r.get("status")),
            "Timestamp": format_timestamp_range(r),
        })
    return rows


def resolve_output_xlsx_path(raw_value: str) -> str:
    """
    Turn the --output-xlsx argument into a concrete file path:
      - "" (flag given with no value) -> current directory + generated name
      - an existing directory         -> that directory + generated name
      - anything else                 -> used as-is (a full file path)
    """
    default_name = f"Daily_QRadar_{datetime.now().strftime('%d%m%Y-%H-%M-%S')}.xlsx"
    if not raw_value:
        return os.path.join(os.getcwd(), default_name)
    if os.path.isdir(raw_value):
        return os.path.join(raw_value, default_name)
    return raw_value


def write_xlsx(rows: list, output_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        raise RuntimeError(
            "--output-xlsx requires the 'openpyxl' package. Install it with: "
            "pip install openpyxl"
        ) from None

    wb = Workbook()
    ws = wb.active
    ws.title = "Offenses"

    ws.append(EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row[col] for col in EXCEL_COLUMNS])

    last_row = max(len(rows) + 1, 1)
    if rows:
        last_col_letter = get_column_letter(len(EXCEL_COLUMNS))
        table = Table(displayName="Offenses", ref=f"A1:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True,
        )
        ws.add_table(table)

    for idx, col in enumerate(EXCEL_COLUMNS, start=1):
        max_len = max([len(col)] + [len(str(row[col])) for row in rows]) if rows else len(col)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    wb.save(output_path)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Fetch QRadar OCSF offenses (GET /siem/offenses_ocsf) by offense "
                     "ID range (finding_info.uid), using position estimation with "
                     "automatic full-scan fallback for correctness.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--host", default=None,
                   help="QRadar console hostname or IP. Required unless using a --kb-* "
                        "management flag.")
    p.add_argument("--start-id", type=int, default=None,
                   help="First offense ID (e.g. 5001). Required unless using a --kb-* "
                        "management flag.")
    p.add_argument("--end-id", type=int, default=None,
                   help="Last offense ID, inclusive. Required unless using a --kb-* "
                        "management flag.")
    p.add_argument("--username", "-u", default=None,
                   help="QRadar username. Required unless using a --kb-* management flag.")
    p.add_argument("--password", "-p", default=None,
                   help="QRadar password. Omit to be prompted securely (recommended).")
    p.add_argument("--yes", "-y", action="store_true",
                   help=f"Skip the confirmation prompt shown when requesting more than "
                        f"{LARGE_REQUEST_THRESHOLD} offense IDs. Required in non-interactive "
                        "sessions if the range is that large.")
    p.add_argument("--api-version", default=DEFAULT_API_VERSION, help="Version header value")
    p.add_argument("--page-size", type=int, default=DEFAULT_PAGE_SIZE,
                   help="Offenses per Range-header page")
    p.add_argument("--buffer", type=int, default=DEFAULT_BUFFER,
                   help="Initial position padding around the estimated window")
    p.add_argument("--max-expand", type=int, default=DEFAULT_MAX_EXPAND,
                   help="Max buffer-doubling retries before falling back to full scan")
    p.add_argument("--full-scan", action="store_true",
                   help="Skip position estimation; scan the entire offense list and "
                        "filter client-side. Slower, but always correct.")
    p.add_argument("--severity", default=None, metavar="LEVEL[,LEVEL...]",
                   help="Only include offenses matching these severities (case-insensitive, "
                        "comma-separated), e.g. 'High,Critical'. Matches the OCSF 'severity' "
                        "string field (or severity_id if severity is absent). Applied after "
                        "the ID-range fetch -- the missing-ID completeness check still "
                        "considers the full unfiltered set.")
    p.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT, help="Per-request timeout")
    p.add_argument("--retries", type=int, default=DEFAULT_RETRIES,
                   help="Retries on HTTP 5xx / connection errors")
    p.add_argument("--insecure", action="store_true",
                   help="Skip TLS certificate verification (lab/self-signed only)")
    p.add_argument("--output-xlsx", nargs="?", const="", default=None, metavar="PATH",
                   help="Write results as an Excel report table instead of printing JSON "
                        "to stdout. PATH may be a full file path, an existing directory "
                        "(default filename is used inside it), or omitted entirely (writes "
                        f"'{DEFAULT_XLSX_NAME_PATTERN}' to the current directory). "
                        "Requires 'pip install openpyxl'.")
    p.add_argument("--kb-db", default=None, metavar="PATH",
                   help="Path to the SQLite knowledge-base file used for Deskripsi/"
                        f"Rekomendasi lookups (default: '{DEFAULT_KB_DB_NAME}' next to "
                        "this script). Auto-created on first use.")
    p.add_argument("--no-kb", action="store_true",
                   help="Disable KB lookup when generating a report; Deskripsi falls back "
                        "to finding_info.desc and Rekomendasi stays 'N/A'.")
    p.add_argument("--kb-import", default=None, metavar="XLSX_PATH",
                   help="[KB management] Bulk import/upsert KB entries from an Excel file "
                        "(columns: ID, Offense Name, Deskripsi, Rekomendasi), then exit. "
                        "Idempotent -- re-importing updates existing rows.")
    p.add_argument("--kb-list", action="store_true",
                   help="[KB management] List all KB entries (id + offense name), then exit.")
    p.add_argument("--kb-add", action="store_true",
                   help="[KB management] Add one KB entry. Requires --kb-keyword, "
                        "--kb-deskripsi, and --kb-rekomendasi.")
    p.add_argument("--kb-update", type=int, default=None, metavar="ID",
                   help="[KB management] Update a KB entry by id. Requires at least one of "
                        "--kb-keyword/--kb-deskripsi/--kb-rekomendasi.")
    p.add_argument("--kb-delete", type=int, default=None, metavar="ID",
                   help="[KB management] Delete a KB entry by id, then exit.")
    p.add_argument("--kb-keyword", default=None, metavar="TEXT",
                   help="Keyword/offense-name text for --kb-add or --kb-update.")
    p.add_argument("--kb-deskripsi", default=None, metavar="TEXT",
                   help="Deskripsi text for --kb-add or --kb-update.")
    p.add_argument("--kb-rekomendasi", default=None, metavar="TEXT",
                   help="Rekomendasi text for --kb-add or --kb-update.")
    return p.parse_args()


def confirm_large_request(start_id: int, end_id: int, skip_confirmation: bool) -> None:
    """Warn and require confirmation before fetching more than
    LARGE_REQUEST_THRESHOLD offense IDs -- a large range can be slow and put
    real load on the QRadar API. Exits the process if not confirmed."""
    requested = end_id - start_id + 1
    if requested <= LARGE_REQUEST_THRESHOLD:
        return

    print(f"[WARN] This request covers {requested} offense IDs ({start_id}-{end_id}), "
          f"which is more than {LARGE_REQUEST_THRESHOLD}. This may be slow and puts "
          "real load on the QRadar API.", file=sys.stderr)

    if skip_confirmation:
        return

    if not sys.stdin.isatty():
        print("[ERROR] Non-interactive session -- re-run with --yes to confirm large "
              "requests.", file=sys.stderr)
        sys.exit(2)

    answer = input(f"Continue fetching {requested} offense IDs? [y/N]: ").strip().lower()
    if answer not in ("y", "yes"):
        print("[INFO] Aborted by user.", file=sys.stderr)
        sys.exit(0)


def main():
    args = parse_args()

    kb_db_path = resolve_kb_db_path(args.kb_db)

    kb_management_requested = any([
        args.kb_import, args.kb_list, args.kb_add,
        args.kb_update is not None, args.kb_delete is not None,
    ])
    if kb_management_requested:
        run_kb_management(args, kb_db_path)
        return

    if not (args.host and args.start_id is not None and args.end_id is not None and args.username):
        print("[ERROR] --host, --start-id, --end-id, and --username/-u are required "
              "unless using a --kb-* management flag.", file=sys.stderr)
        sys.exit(2)

    if args.start_id > args.end_id:
        print("[ERROR] --start-id must be <= --end-id", file=sys.stderr)
        sys.exit(2)

    confirm_large_request(args.start_id, args.end_id, args.yes)

    password = args.password
    if password is None:
        password = getpass.getpass(f"QRadar password for {args.username}: ")
    elif sys.stdin.isatty():
        print("[WARN] --password on the command line can leak via shell history/`ps`. "
              "Prefer the secure prompt.", file=sys.stderr)

    headers = {
        "Authorization": build_basic_auth_header(args.username, password),
        "Version": args.api_version,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    del password

    if args.insecure:
        print("[WARN] --insecure: TLS certificate verification is disabled.", file=sys.stderr)
        ssl_context = ssl._create_unverified_context()
    else:
        ssl_context = ssl.create_default_context()

    try:
        if args.full_scan:
            matches = full_scan(args.host, args.start_id, args.end_id, headers,
                                 ssl_context, args.timeout, args.retries, args.page_size)
        else:
            matches, verified = smart_fetch(args.host, args.start_id, args.end_id, headers,
                                             ssl_context, args.timeout, args.retries,
                                             args.page_size, args.buffer, args.max_expand)
            if not verified:
                print("[WARN] Could not verify the position-estimated window was complete "
                      "(likely ID gaps or non-ID-ordered list). Falling back to a full "
                      "scan for a correctness guarantee.", file=sys.stderr)
                matches = full_scan(args.host, args.start_id, args.end_id, headers,
                                     ssl_context, args.timeout, args.retries, args.page_size)
    except RuntimeError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)

    target_ids = set(range(args.start_id, args.end_id + 1))
    found_ids = {u for r in matches if (u := get_uid(r)) is not None}
    missing_ids = sorted(target_ids - found_ids)
    if missing_ids:
        print(f"[WARN] {len(missing_ids)} offense ID(s) not found (deleted, purged, or "
              f"out of this account's scope): {missing_ids}", file=sys.stderr)

    matches.sort(key=lambda r: get_uid(r) or 0)

    if args.severity:
        allowed = parse_severity_filter(args.severity)
        before = len(matches)
        matches = filter_by_severity(matches, allowed)
        print(f"[INFO] --severity filter ({args.severity}): {len(matches)}/{before} "
              f"offense(s) kept.", file=sys.stderr)

    if args.output_xlsx is not None:
        output_path = resolve_output_xlsx_path(args.output_xlsx)
        kb_conn = None
        try:
            if not args.no_kb:
                kb_conn = kb_connect(kb_db_path)
                if kb_entry_count(kb_conn) == 0:
                    print(f"[INFO] KB database at {kb_db_path} is empty -- Deskripsi/"
                          "Rekomendasi will use fallback values. Run --kb-import to "
                          "populate it.", file=sys.stderr)
            rows = transform_to_rows(matches, kb_conn)
            write_xlsx(rows, output_path)
        except RuntimeError as e:
            print(f"[ERROR] {e}", file=sys.stderr)
            sys.exit(1)
        finally:
            if kb_conn is not None:
                kb_conn.close()
        print(f"[INFO] Wrote {len(matches)} offense(s) to {output_path}", file=sys.stderr)
    else:
        print(json.dumps(matches, indent=2))


if __name__ == "__main__":
    main()
